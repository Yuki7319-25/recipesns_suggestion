import random
import pandas as pd
import os
from django.shortcuts import render
from django.views.generic import CreateView, TemplateView
from django.urls import reverse_lazy
from .forms import CustomUserCreationForm
from .food import Wasyoku, Yousyoku, Chuuka, Azia, Chuutou, Itiran
from openpyxl import load_workbook
from django.conf import settings
import os
from django.conf import settings
import openpyxl
import random

file_path = os.path.join(settings.BASE_DIR, 'ryouri.xlsx')
book = openpyxl.load_workbook(file_path)


def today_menu_view(request):
    import openpyxl
    import random
    import os
    from .food import Wasyoku, Yousyoku, Chuuka, Azia, Chuutou, Itiran
    from django.conf import settings

    file_path = os.path.join(settings.BASE_DIR, 'ryouri.xlsx')
    book = openpyxl.load_workbook(file_path)
    sheet = book.active

    # 料理名リスト
    c_list = [row[2].value for row in sheet.iter_rows(min_row=2, max_row=16)]

    # GET の場合
    if request.method == "GET":
        sele = random.choice(c_list)
        return render(request, "accounts/today_menu.html", {
            "sele": sele,
            "message": None,
            "result": None,
            "total": None,
        })

    # POST の場合
    sele = request.POST.get("sele")
    ans = request.POST.get("answer")

    if ans not in ["はい", "いいえ"]:
        return render(request, "accounts/today_menu.html", {
            "sele": sele,
            "error": "はいかいいえで答えてください",
            "message": None,
            "result": None,
            "total": None,
        })

    if ans == "はい":
        obj = None
        for row in sheet.iter_rows(min_row=2, max_row=16):
            if row[2].value == sele:
                genre = row[0].value
                country = row[1].value
                staple = row[5].value
                extra = row[4].value

                if genre == "和食":
                    obj = Wasyoku(sele, country, staple, extra)
                elif genre == "洋食":
                    obj = Yousyoku(sele, country, staple, extra)
                elif genre == "中華":
                    obj = Chuuka(sele, country, staple, extra)
                elif genre == "アジア":
                    obj = Azia(sele, country, staple, extra)
                elif genre == "中東":
                    obj = Chuutou(sele, country, staple, extra)
                else:
                    obj = Itiran(sele, country, staple)

                # カウント加算
                count_cell = row[3]
                current = count_cell.value if isinstance(count_cell.value, int) else 0
                count_cell.value = current + 1
                break

        # 合計
        total = sum([row[3].value if isinstance(row[3].value, int) else 0
                     for row in sheet.iter_rows(min_row=2, max_row=16)])
        sheet["A17"] = "集計"
        sheet["D17"] = total
        book.save(file_path)

        # obj が None の場合でも安全に表示
        message = obj.show() if obj else "情報が取得できませんでした"

        return render(request, "accounts/today_menu.html", {
            "sele": sele,
            "message": message,
            "result": None,
            "total": total,
        })

    # 「いいえ」の場合
    return render(request, "accounts/today_menu.html", {
        "sele": sele,
        "message": "左様ですか",
        "result": None,
        "total": None,
    })




# ------------------------------
# 既存ビュー
# ------------------------------
def calc_view(request):
    from scripts.calc import add_numbers
    result = add_numbers(5, 7)
    return render(request, "accounts/today_menu.html", {"result": result})


class IndexView(TemplateView):
    template_name = 'index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["welcome_message"] = "ようこそ！"

        # スライドショーに使う画像をここに書く
        context["slider_images"] = [
            "/static/photo/images/mizu.jpg",
            "/static/photo/images/wood.jpg",
        ]

        return context




class SignUpView(CreateView):
    form_class = CustomUserCreationForm
    template_name = "accounts/signup.html"
    success_url = reverse_lazy('accounts:signup_success')


class SignUpSuccessView(TemplateView):
    template_name = "accounts/signup_success.html"

def welcom_boad(request):
    return render(request, 'accounts/welcom_boad.html')

from django.contrib.auth.decorators import login_required

@login_required  # ログインユーザーのみ閲覧可能
def mypage_view(request):
    # 必要に応じてユーザー情報や投稿データを渡す
    context = {
        "user": request.user,
    }
    return render(request, "photo/mypage.html", context)
